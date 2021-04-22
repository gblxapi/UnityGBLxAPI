# -------------------------------------------------------------------------------------------------
# GBLxAPI_Json_Parser.py
# Project: GBLXAPI
# Created: 2018/07/21
# Copyright 2018 Dig-It! Games, LLC. All rights reserved.
# This code is licensed under the MIT License (See LICENSE.txt for details)
# -------------------------------------------------------------------------------------------------

from openpyxl import load_workbook
import json
from jsonmerge import merge

# This function takes all of the GBLxAPI Vocabulary information in the workbook named workbookName
# and parses it to json, writing to a file with the name defined in target.
def GenerateJson(workbookName, target, nameCol, uriCol, descrCol):
    wb = load_workbook(filename=workbookName)

    totalMap = {} # totalMap has keys in [Activity, Grade, Domain, Focus, etc]
    for ws in wb:
        wsName = ws._WorkbookChild__title

        if wsName == "Notes": continue

        print("Loading " + wsName +"..."),

        sectionMap = {} # sectionMap has keys in [Counting, Algebra, Energy, etc]

        # local variables to allow for column overrides
        nc = nameCol
        uc = uriCol
        dc = descrCol

        # override column values for specific manually populated sheets in the default file
        # for automatically populated sheets, the default file uses columns F, I, and BB. For manual population, it's much easier to use A, B, and C.
        # This should not affect the values for the user vocab, since this file uses A, B, and C already.
        if wsName in ["Verb", "Activity", "Extension", "Grade"]:
            nc = 0 # A
            uc = 1 # B
            dc = 2 # C

        for row in ws.iter_rows(min_row=2): # min_row=2 to skip header row
            itemMap = {} # itemMap has keys in [name, description, id]

            # force all values to lowercase for easy comparison
            name = str(row[nc].value).lower() if row[nc].value is not None else ""
            uri = str(row[uc].value).lower() if row[uc].value is not None else ""
            descr = str(row[dc].value).lower() if row[dc].value is not None else ""

            # populate the map with the corresponding values
            itemMap['name'] = {}
            itemMap['description'] = {}

            itemMap['name']['en-US'] = name
            itemMap['id'] = uri
            itemMap['description']['en-US'] = descr

            sectionMap[name] = itemMap

        totalMap[wsName.lower()] = sectionMap

        print("Done.")

    print("Generating Json file..."),
    with open(target, 'w') as write_file:
        json.dump(totalMap, write_file, sort_keys=True, indent=4, separators=(',', ': '))
        print("Success!")

print("Converting your data...")

# Load default vocabulary
# 5 == row F, 8 == row I, 53 == row BB in Excel
print("Loading default vocabulary...")
GenerateJson('GBLxAPI_Vocab_Default.xlsx', 'GBLxAPI_Vocab_Default.json', 5, 53, 8)

# Load user overrides
# 0 == row A, 1 == row B, 2 == row C in Excel
print("Loading user overrides...")
GenerateJson('GBLxAPI_Vocab_User.xlsx', 'GBLxAPI_Vocab_User.json', 0, 1, 2)

print("All done! Move the two generated Json files to Resources/Data to use the GBLxAPI vocabulary in your Unity project.")